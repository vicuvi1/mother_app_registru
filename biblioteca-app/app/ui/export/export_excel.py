"""Export Excel cu openpyxl — pagini stivuite, grilă completă, anteturi de grup, numerotare la print."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from ui.export.export_html import group_spans
from ui.export.export_utils import format_cell_value, format_total_value, validate_pages

_thin = Side(style="thin", color="555555")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BOLD = Font(bold=True, size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_FILL = PatternFill("solid", fgColor="EEF2F7")
GROUP_FILL = PatternFill("solid", fgColor="DBE3EE")
TOTAL_FILL = PatternFill("solid", fgColor="E2E8F0")


def export_to_excel(out_path: Path, pages: list[dict]) -> Path:
    validate_pages(pages)
    wb = Workbook()
    ws = wb.active
    ws.title = "Registru"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.right.text = "Pagina &P din &N"
    ws.evenFooter.right.text = "Pagina &P din &N"

    max_cols = max((len(p["headers"]) for p in pages if p.get("type") != "cover"), default=1)

    r = 1
    for pi, page in enumerate(pages):
        if pi > 0:
            ws.row_breaks.append(Break(id=r - 1))
        if page.get("type") == "cover":
            r = _write_cover(ws, page, r, max_cols)
        else:
            r = _write_page(ws, page, r)
        r += 2  # spațiu între pagini

    for c in range(1, max_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.column_dimensions["A"].width = 13

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _write_cover(ws, page: dict, r: int, ncols: int) -> int:
    r += 6
    lines = [
        (page.get("institutie_1"), 14),
        (page.get("institutie_2"), 14),
        (None, 0),
        (page.get("titlu"), 22),
        (page.get("biblioteca"), 18),
        (None, 0),
        (page.get("localitate"), 14),
        (page.get("an"), 16),
    ]
    for text, size in lines:
        if not text:
            r += 1
            continue
        cell = ws.cell(r, 1, text)
        cell.font = Font(bold=True, size=size)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if ncols > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        r += 1
    return r


def _write_page(ws, page: dict, r: int) -> int:
    headers = page["headers"]
    groups = page["groups"]
    col_keys = page["col_keys"]
    rows = page["rows"]
    total_rows = page["total_rows"]
    meta = page["meta"]
    ncols = len(headers)

    title_lines = []
    if meta.get("nume_biblioteca"):
        loc = f", {meta['localitate']}" if meta.get("localitate") else ""
        title_lines.append(f"{meta['nume_biblioteca']}{loc}")
    title_lines.append("Registru de evidență a activității bibliotecii")
    partea = f"Partea {meta.get('parte_roman', '')}. {meta.get('title', '')}"
    if meta.get("luna_name"):
        partea += f" în luna {meta['luna_name']} anul {meta.get('an', '')}"
    elif meta.get("an"):
        partea += f" — anul {meta.get('an', '')}"
    title_lines.append(partea)

    for line in title_lines:
        ws.cell(r, 1, line).font = Font(bold=True, size=11)
        if ncols > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        ws.cell(r, 1).alignment = Alignment(horizontal="center")
        r += 1

    has_groups = any(groups)
    spans = group_spans(groups)

    if has_groups:
        for start, end, g in spans:
            c0, c1 = start + 1, end + 1
            if g:
                ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c1)
                cell = ws.cell(r, c0, g)
                cell.font = BOLD
                cell.alignment = CENTER
                cell.fill = GROUP_FILL
                for cc in range(c0, c1 + 1):
                    ws.cell(r, cc).border = BORDER
            else:
                ws.merge_cells(start_row=r, start_column=c0, end_row=r + 1, end_column=c0)
                cell = ws.cell(r, c0, headers[start])
                cell.font = BOLD
                cell.alignment = CENTER
                cell.fill = HEADER_FILL
                cell.border = BORDER
        r += 1

    for c, h in enumerate(headers, 1):
        grp = groups[c - 1] if c - 1 < len(groups) else ""
        if has_groups and not grp:
            continue
        cell = ws.cell(r, c, h)
        cell.font = BOLD
        cell.alignment = CENTER
        cell.fill = HEADER_FILL
        cell.border = BORDER
    r += 1

    for row in rows:
        for c, key in enumerate(col_keys, 1):
            val = format_cell_value(row.get(key, ""))
            cell = ws.cell(r, c, val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c == 1:
                cell.font = BOLD
        r += 1

    for label, sums in total_rows:
        cell = ws.cell(r, 1, label)
        cell.font = BOLD
        cell.fill = TOTAL_FILL
        cell.border = BORDER
        for c, key in enumerate(col_keys, 1):
            if c == 1:
                continue
            val = format_total_value(sums.get(key, ""))
            cell = ws.cell(r, c, val)
            cell.font = BOLD
            cell.fill = TOTAL_FILL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
        r += 1

    return r
