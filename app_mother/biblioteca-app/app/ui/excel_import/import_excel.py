"""Import rânduri din fișier Excel exportat de aplicație."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import and_, delete

from core.constants_manager import get_all_etichete
from core.part_import_meta import get_part_import_meta
from database.db_manager import get_session
from ui.widgets.table.column_def import ColumnDef


def _coerce_value(col: ColumnDef, raw: Any) -> Any:
    if raw is None:
        raw = ""
    if col.col_type == "int":
        try:
            return max(0, int(float(raw))) if str(raw).strip() else 0
        except (TypeError, ValueError):
            return 0
    if col.col_type in ("bool",) or col.col_type.startswith("scope_"):
        text = str(raw).strip().casefold()
        return text in ("1", "da", "true", "x", "✓")
    if col.col_type == "date":
        if hasattr(raw, "strftime"):
            return raw.strftime("%Y-%m-%d")
        return str(raw).strip()
    return str(raw).strip() if raw is not None else ""


def parse_excel_rows(path: Path, part_id: str) -> list[dict[str, Any]]:
    meta = get_part_import_meta(part_id)
    columns: list[ColumnDef] = meta["columns"]
    labels = get_all_etichete(part_id)
    header_texts = [labels.get(c.key, c.key) for c in columns]

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    sheet_rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        if row and any(c is not None and str(c).strip() for c in row):
            sheet_rows.append(list(row))

    header_idx = _find_header_row(sheet_rows, header_texts)
    if header_idx is None:
        raise ValueError(
            "Nu s-a găsit antetul coloanelor în Excel. "
            "Folosiți un fișier exportat din această aplicație."
        )

    data_rows: list[dict[str, Any]] = []
    for row in sheet_rows[header_idx + 1 :]:
        first = str(row[0] or "").strip().casefold()
        if first.startswith("total"):
            break
        if not any(c is not None and str(c).strip() for c in row):
            continue
        item: dict[str, Any] = {}
        for j, col in enumerate(columns):
            if j >= len(row):
                break
            item[col.key] = _coerce_value(col, row[j])
        data_rows.append(item)
    return data_rows


def _find_header_row(sheet_rows: list[list[Any]], header_texts: list[str]) -> int | None:
    n = len(header_texts)
    for i, row in enumerate(sheet_rows):
        cells = [str(row[j] or "").strip() if j < len(row) else "" for j in range(n)]
        if cells == header_texts:
            return i
        if sum(1 for j in range(n) if cells[j] == header_texts[j]) >= max(2, n // 2):
            return i
    return None


def import_rows_to_database(
    part_id: str,
    year: int,
    month: int | None,
    category: str | None,
    rows: list[dict[str, Any]],
    *,
    replace: bool = True,
) -> int:
    meta = get_part_import_meta(part_id)
    model = meta["model"]
    mode = meta["mode"]
    columns: list[ColumnDef] = meta["columns"]
    model_keys = {c.name for c in model.__table__.columns}

    with get_session() as session:
        if replace:
            filters = []
            if hasattr(model, "an"):
                filters.append(model.an == year)
            if month is not None and hasattr(model, "luna") and mode in ("daily", "events"):
                filters.append(model.luna == month)
            if category and hasattr(model, "categorie_varsta"):
                filters.append(model.categorie_varsta == category)
            if filters:
                session.execute(delete(model).where(and_(*filters)))

        count = 0
        for i, row in enumerate(rows):
            kwargs: dict[str, Any] = {}
            if hasattr(model, "an"):
                kwargs["an"] = year
            if month is not None and hasattr(model, "luna") and mode in ("daily", "events"):
                kwargs["luna"] = month
            if mode == "monthly" and "luna" in row:
                kwargs["luna"] = int(row["luna"])
            if category and hasattr(model, "categorie_varsta"):
                kwargs["categorie_varsta"] = category
            for col in columns:
                if col.key in row and col.key in model_keys:
                    kwargs[col.key] = row[col.key]
            if mode == "events" and meta["date_field"] in model_keys:
                if not kwargs.get(meta["date_field"]):
                    kwargs[meta["date_field"]] = f"_imp{i + 1}"
            kwargs = {k: v for k, v in kwargs.items() if k in model_keys}
            if "is_auto_generated" in model_keys:
                kwargs["is_auto_generated"] = False
            session.add(model(**kwargs))
            count += 1
        session.commit()
    return count
